from flask import Flask, jsonify
import openpyxl
import os

def create_app(inventory_excel_path: str = "data/inventario_ejemplo.xlsx",
               sales_excel_path: str = "data/ventas_ejemplo.xlsx"):
    app = Flask(__name__)

    @app.route("/inventario")
    def inventario():
        if not os.path.exists(inventory_excel_path):
            return jsonify({"error": "Archivo no encontrado"}), 404
        wb = openpyxl.load_workbook(inventory_excel_path)
        ws = wb["Resumen"]
        data = []
        for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
            if row and row[0] and row[4]:
                data.append({"estilo": row[0], "litros_total": row[4], "cop": row[5]})
        wb.close()
        return jsonify(data)

    @app.route("/inventario/fermentadores")
    def fermentadores():
        if not os.path.exists(inventory_excel_path):
            return jsonify({"error": "Archivo no encontrado"}), 404
        wb = openpyxl.load_workbook(inventory_excel_path)
        ws = wb["Fermentadores"]
        data = []
        for row in ws.iter_rows(min_row=4, max_row=10, values_only=True):
            if row and row[0]:
                data.append({
                    "fermentador": row[0],
                    "litros": row[2],
                    "estilo": row[4],
                    "alarma": row[5]
                })
        wb.close()
        return jsonify(data)

    @app.route("/ventas/resumen")
    def ventas_resumen():
        if not os.path.exists(sales_excel_path):
            return jsonify({"error": "Archivo no encontrado"}), 404
        wb = openpyxl.load_workbook(sales_excel_path)
        ws = wb["RESUMEN"]
        data = []
        for row in ws.iter_rows(min_row=4, max_row=14, values_only=True):
            if row and row[0]:
                data.append({"mes": row[0], "facturas": row[1], "total": row[2]})
        wb.close()
        return jsonify(data)

    return app

