import time
import openpyxl
import os
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, List, Optional, Tuple
import logging
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Z1Agent:
    # Mapeo de productos
    PRODUCT_MAP = {
        "PTB01": {"id": "XX", "name": "Golden Ale 330ml", "price": 8500, "style": "BLONDE ALE"},
        "PTB02": {"id": "XX", "name": "Irish Red Ale 330ml", "price": 8500, "style": "IRISH RED ALE"},
        "PTB03": {"id": "XX", "name": "APA 330ml", "price": 8500, "style": "APA"},
        "PTB04": {"id": "XX", "name": "IPA 330ml", "price": 9500, "style": "IPA"},
        "PTB05": {"id": "XX", "name": "Stout 330ml", "price": 8500, "style": "STOUT"},
    }

    def __init__(self, sheet_path: str, alegra_client, inventory_manager):
        self.sheet_path = sheet_path
        self.alegra = alegra_client
        self.inventory = inventory_manager

    def _extract_code(self, value: str) -> Optional[str]:
        import re
        if not isinstance(value, str):
            return None
        match = re.search(r'PTB\d{2}', value.upper())
        return match.group(0) if match else None

    def _should_process(self, row: tuple, headers: list) -> Tuple[bool, Optional[str]]:
        try:
            facturar_index = headers.index("Facturar: Sí") if "Facturar: Sí" in headers else -1
            factura_index = headers.index("Factura") if "Factura" in headers else -1
        except ValueError:
            return False, None

        debe_facturar = row[facturar_index] if facturar_index > -1 and len(row) > facturar_index else False
        tiene_factura = row[factura_index] if factura_index > -1 and len(row) > factura_index else False

        if debe_facturar == "Sí" and not tiene_factura:
            return True, None
        return False, None

    def run(self) -> Dict:
        start = time.time()
        facturas_creadas = 0
        productos_vendidos = defaultdict(float)
        errors = []

        if not os.path.exists(self.sheet_path):
            return {"ok": False, "error": f"Archivo no encontrado: {self.sheet_path}"}

        wb = openpyxl.load_workbook(self.sheet_path)
        ws = wb["Respuestas de formulario 1"]
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]

        col_cliente = headers.index("Nombre del cliente") if "Nombre del cliente" in headers else -1
        col_productos = [headers.index(col) for col in headers if "Producto" in col][:5]
        col_cantidades = [headers.index(col) for col in headers if "Cantidad" in col][:5]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            should_process, _ = self._should_process(row, headers)
            if not should_process:
                continue

            cliente_nombre = row[col_cliente] if col_cliente > -1 and len(row) > col_cliente else None
            if not cliente_nombre:
                continue

            client = self.alegra.search_client_by_name(cliente_nombre)
            if not client:
                errors.append(f"Cliente no encontrado: {cliente_nombre}")
                continue

            items = []
            for prod_col, cant_col in zip(col_productos, col_cantidades):
                if prod_col > -1 and cant_col > -1 and len(row) > prod_col and len(row) > cant_col:
                    ref = self._extract_code(row[prod_col])
                    cantidad = float(row[cant_col]) if row[cant_col] else 0
                    if ref and cantidad > 0 and ref in self.PRODUCT_MAP:
                        product = self.PRODUCT_MAP[ref]
                        items.append({
                            "id": product["id"],
                            "quantity": cantidad,
                            "price": product["price"]
                        })
                        productos_vendidos[product["style"]] += cantidad

            if not items:
                continue

            due_date = (datetime.now().date() + timedelta(days=30)).isoformat()
            factura = self.alegra.create_invoice(client["id"], items, due_date)

            if "error" in factura:
                errors.append(f"Error al crear factura para {cliente_nombre}: {factura['error']}")
                continue

            facturas_creadas += 1
            factura_index = headers.index("Factura") if "Factura" in headers else -1
            if factura_index > -1:
                ws.cell(row=row_idx, column=factura_index+1, value=f"Sí - Factura {factura.get('id', '')}")

        stock_actualizado = {}
        if productos_vendidos:
            stock_actualizado = self.inventory.deduct_stock(dict(productos_vendidos))

        wb.save(self.sheet_path)

        return {
            "ok": True,
            "facturas_creadas": facturas_creadas,
            "productos_vendidos": dict(productos_vendidos),
            "stock_actualizado": stock_actualizado,
            "errors": errors,
            "duracion_seg": round(time.time() - start, 2)
        }

